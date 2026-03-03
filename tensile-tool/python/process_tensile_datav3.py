#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
拉伸试验数据处理脚本 v3.0
依赖：numpy, openpyxl（无 pandas）
"""

import os, re, sys, math, csv, warnings
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

EXT_RANGES = {10: 0.2, 15: 0.3, 25: 0.5}

def _read_raw_csv(filepath):
    for enc in ('gbk', 'gb18030', 'utf-8-sig', 'utf-8', 'latin-1'):
        try:
            with open(filepath, encoding=enc, newline='') as f:
                rows = [
                    [c.strip('"').strip() for c in row]
                    for row in csv.reader(f)
                    if any(c.strip('"').strip() for c in row)
                ]
            print(f"  -> 文件编码：{enc}")
            return rows
        except UnicodeDecodeError:
            continue
    raise IOError(f"无法解码文件：{filepath}")

def detect_specimens(rows):
    row0 = rows[0]
    specimens, col = [], 0
    while col < len(row0):
        name = row0[col].strip()
        if name and name.lower() != 'nan':
            specimens.append((name, col))
            col += 4
        else:
            col += 1
    return specimens

def extract_specimen(rows, col_start):
    data = {'时间': [], '载荷': [], '行程': [], '引伸计': []}
    keys = list(data.keys())
    for row in rows[3:]:
        for ki, key in enumerate(keys):
            ci = col_start + ki
            data[key].append(row[ci] if ci < len(row) else '')
    return data

def find_csv_files(start_file):
    folder = os.path.dirname(os.path.abspath(start_file))
    name, ext = os.path.splitext(os.path.basename(start_file))
    m = re.match(r'^(.+)-(\d+)$', name)
    if not m:
        return [start_file]
    prefix  = m.group(1)
    pattern = re.compile(rf'^{re.escape(prefix)}-(\d+){re.escape(ext)}$', re.I)
    found   = {}
    for fn in os.listdir(folder):
        pm = pattern.match(fn)
        if pm:
            found[int(pm.group(1))] = os.path.join(folder, fn)
    if not found:
        raise FileNotFoundError(f"未找到 {prefix}-*.csv")
    files = [found[k] for k in sorted(found)]
    print(f"\n  发现 {len(files)} 个分段文件：" + ', '.join(os.path.basename(f) for f in files))
    return files

def load_and_merge(files):
    all_rows = [_read_raw_csv(f) for f in files]
    specimens = detect_specimens(all_rows[0])
    if len(files) == 1:
        return all_rows[0], specimens
    header = all_rows[0][:3]
    data   = [r for rows in all_rows for r in rows[3:]]
    return header + data, specimens

def _safe_path(desired):
    if not os.path.exists(desired):
        return desired
    if input(f"  ⚠ 已存在：{desired}，覆盖？[y/N]: ").strip().lower() == 'y':
        return desired
    base, ext = os.path.splitext(re.sub(r'_\d+$', '', os.path.splitext(desired)[0]))
    ext = os.path.splitext(desired)[1]
    i = 1
    while os.path.exists(f"{base}_{i}{ext}"):
        i += 1
    path = f"{base}_{i}{ext}"
    print(f"  -> 另存为：{path}")
    return path

def _write_safe(fn, desired):
    path = _safe_path(desired)
    while True:
        try:
            fn(path); return path
        except PermissionError:
            base = re.sub(r'_\d+$', '', os.path.splitext(path)[0])
            ext  = os.path.splitext(path)[1]
            i = 1
            while os.path.exists(f"{base}_{i}{ext}"):
                i += 1
            path = f"{base}_{i}{ext}"
            print(f"  ✗ 文件被占用，另存为：{path}")

def save_pre_xlsx(rows, specimens, prefix):
    def _write(path):
        wb = Workbook()
        wb.remove(wb.active)
        for sp_name, col_start in specimens:
            col_labels = [rows[1][col_start+i] if col_start+i < len(rows[1]) else '' for i in range(4)]
            units      = [rows[2][col_start+i] if col_start+i < len(rows[2]) else '' for i in range(4)]
            ws = wb.create_sheet(sp_name[:31])
            for ci, lbl in enumerate(col_labels, 1):
                ws.cell(1, ci, lbl)
            for ci, u in enumerate(units, 1):
                c = ws.cell(2, ci, u)
                c.font = Font(italic=True, color='777777')
            for ri, row in enumerate(rows[3:], 3):
                for ci in range(4):
                    ws.cell(ri, ci+1, row[col_start+ci] if col_start+ci < len(row) else '')
        wb.save(path)
    out = _write_safe(_write, f"pre-{prefix}.xlsx")
    print(f"  -> pre 文件：{out}")

def _calc_area(w, t, cyl):
    if cyl:
        a = w * t * math.pi / 4
        return a, f"圆柱棒 π/4×{w}²={a:.4f} mm²"
    a = w * t
    return a, f"矩形 {w}×{t}={a:.4f} mm²"

def _input_one(name):
    print(f"\n  【{name}】")
    while True:
        try:
            w  = float(input("    宽度/直径 (mm): "))
            tr = input("    厚度 (mm)（圆柱棒直接回车）: ").strip()
            cyl = tr == ''
            t   = w if cyl else float(tr)
            g   = float(input("    标距 (mm): "))
            break
        except ValueError:
            print("    ✗ 请输入数字")
    if not cyl and abs(w - t) < 1e-9:
        cyl = input(f"    宽度=厚度={w}，圆柱棒？[y/N]: ").strip().lower() == 'y'
    a, desc = _calc_area(w, t, cyl)
    print(f"    -> {desc}，标距={g} mm")
    return {'width': w, 'thickness': t, 'gauge': g, 'area': a}

def _gen_template(specimens, path):
    wb = Workbook()
    ws = wb.active
    ws.title = '试样参数'
    hdrs  = ['试样名称', '宽度/直径 (mm)', '厚度 (mm)', '标距 (mm)', '是否圆柱棒(是/否)']
    fills = ['C6EFCE', 'FFEB9C', 'FFEB9C', 'FFEB9C', 'DDEBF7']
    for ci, (h, fc) in enumerate(zip(hdrs, fills), 1):
        c = ws.cell(1, ci, h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
        c.fill = PatternFill(start_color=fc, end_color=fc, fill_type='solid')
    notes = ['<- 勿修改', '<- 矩形填宽，圆柱填直径', '<- 矩形填厚，圆柱留空',
             '<- 引伸计标距', '<- 填"是"按 π/4·d² 算']
    for ci, n in enumerate(notes, 1):
        ws.cell(2, ci, n).font = Font(italic=True, color='999999', size=9)
    for ri, (sp, _) in enumerate(specimens, 3):
        ws.cell(ri, 1, sp).font = Font(bold=True)
        ws.cell(ri, 5, '否')
    for ci, w in enumerate([20, 18, 14, 14, 22], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    wb.save(path)

def _load_template(path, specimens):
    wb  = load_workbook(path, data_only=True)
    ws  = wb.active
    sp_names = [s[0] for s in specimens]
    params = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if name not in sp_names:
            continue
        try:
            w = float(row[1]); g = float(row[3])
        except (TypeError, ValueError) as e:
            print(f"  ✗ {name} 参数错误：{e}"); sys.exit(1)
        tr  = row[2]
        cyl = (tr is None or str(tr).strip() == '' or
               str(row[4]).strip() in ('是','yes','Yes','YES','y','Y','1','true','True'))
        t   = w if (tr is None or str(tr).strip() == '') else float(tr)
        if not cyl and abs(w - t) < 1e-9:
            cyl = True
        a, desc = _calc_area(w, t, cyl)
        print(f"    {name}：{desc}，标距={g} mm")
        params[name] = {'width': w, 'thickness': t, 'gauge': g, 'area': a}
    miss = [s for s in sp_names if s not in params]
    if miss:
        print(f"  ✗ 模板中缺少：{miss}"); sys.exit(1)
    return params

def get_all_params(specimens):
    if len(specimens) == 1:
        return {specimens[0][0]: _input_one(specimens[0][0])}
    tpl = os.path.abspath('specimen_params_template.xlsx')
    print(f"\n── 多试样参数（{len(specimens)} 个）──")
    print("  [1] 生成模板后填写导入（推荐）")
    print("  [2] 指定已有参数文件")
    print("  [3] 逐个手动输入")
    c = input("  选项 [1/2/3]，默认1：").strip() or '1'
    if c == '3':
        return {n: _input_one(n) for n, _ in specimens}
    if c == '2':
        p = input("  参数文件路径：").strip().strip('"')
        if not os.path.exists(p):
            print("  ✗ 文件不存在"); sys.exit(1)
    else:
        _gen_template(specimens, tpl)
        print(f"\n  模板：{tpl}")
        print("  填写宽度/直径、厚度、标距、是否圆柱棒后保存，按回车继续...")
        input()
        p = tpl
    return _load_template(p, specimens)

def _pf(s):
    s = str(s).strip()
    if s in ('-.----', '', 'nan', '-'):
        return float('nan')
    try:
        return float(s)
    except ValueError:
        return float('nan')

def process_extensometer(data, gauge, sp_name):
    print(f"\n  [引伸计] {sp_name}")
    raw_ext = data['引伸计']
    ext_num = [_pf(v) for v in raw_ext]

    first_valid = next((v for v in ext_num if not math.isnan(v)), float('nan'))
    if math.isnan(first_valid) or first_valid == 0.0:
        if not any(not math.isnan(v) and v != 0.0 for v in ext_num):
            if input("  ⚠ 引伸计无效，继续？[y/N]: ").strip().lower() != 'y':
                sys.exit(0)
            data['ext_corrected'] = [float('nan')] * len(raw_ext)
            data['ext_mm'] = 10.0
            return data

    corrected = list(ext_num)
    cum_off, prev, slips = 0.0, None, 0
    for i, v in enumerate(ext_num):
        if math.isnan(v):
            continue
        if prev is not None and (v - prev) < -0.05:
            cum_off += prev - v
            slips   += 1
            print(f"  ⚠ 滑脱 第{i+1}行：{prev:.4f}->{v:.4f} 偏移={cum_off:.4f}")
        corrected[i] = v + cum_off
        prev = v
    if slips:
        print(f"  共 {slips} 次滑脱，已修正")

    dash_idx = next((i for i, v in enumerate(raw_ext) if str(v).strip() == '-.----'), None)
    ext_mm   = None

    if dash_idx is not None and dash_idx > 0:
        last_i   = max(i for i in range(dash_idx) if not math.isnan(corrected[i]))
        last_val = corrected[last_i]
        ext_mm   = min(EXT_RANGES, key=lambda mm: abs(last_val - EXT_RANGES[mm]))
        print(f"  -> {ext_mm}mm 引伸计（末值={last_val:.4f}，满量程≈{EXT_RANGES[ext_mm]}）")
        try:
            last_stroke = _pf(data['行程'][last_i])
        except (IndexError, ValueError):
            last_stroke = 0.0
        print(f"  -> 行程替代：(行程-{last_stroke:.4f})/{gauge}×{ext_mm}+{last_val:.4f}")
        for i in range(dash_idx, len(corrected)):
            if math.isnan(corrected[i]):
                s = _pf(data['行程'][i])
                if not math.isnan(s):
                    corrected[i] = (s - last_stroke) / gauge * ext_mm + last_val

    data['ext_corrected'] = corrected
    data['ext_mm']        = float(ext_mm) if ext_mm else 10.0
    return data

def calculate_stress_strain(data, area):
    ext_mm = data['ext_mm']
    data['load_kN']   = [_pf(v) for v in data['载荷']]
    data['stroke_mm'] = [_pf(v) for v in data['行程']]
    data['stress']    = [v / area * 1000 if not math.isnan(v) else float('nan')
                         for v in data['load_kN']]
    data['strain']    = [v / ext_mm * 100 if not math.isnan(v) else float('nan')
                         for v in data['ext_corrected']]
    return data

def calc_mech_props(data):
    nan = float('nan')
    stress = np.array([v for v in data['stress'] if not math.isnan(v)])
    strain = np.array([v for v in data['strain'] if not math.isnan(v)])
    n = min(len(stress), len(strain))
    stress, strain = stress[:n], strain[:n]
    if n < 10:
        return {'E': nan, 'Rp02': nan, 'Rm': nan, 'A': nan}

    Rm = float(np.max(stress))
    A  = float(np.max(strain))
    E_GPa, Rp02 = nan, nan

    try:
        sd   = strain / 100.0
        idx  = np.where((stress >= 0.05*Rm) & (stress <= 0.70*Rm))[0]
        bE, bcf, bend = nan, None, 5

        if len(idx) >= 5:
            sa, sta = sd[idx], stress[idx]
            for end in range(5, len(idx)+1):
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore')
                    cf = np.polyfit(sa[:end], sta[:end], 1)
                fit = np.polyval(cf, sa[:end])
                sst = np.sum((sta[:end] - np.mean(sta[:end]))**2)
                sse = np.sum((sta[:end] - fit)**2)
                r2  = 1.0 - sse/sst if sst > 1e-12 else 1.0
                if r2 >= 0.9995:
                    bE, bcf, bend = cf[0], cf, end
                else:
                    break

        if math.isnan(bE) and len(idx) >= 5:
            hi = max(10, int(len(idx)*0.2))
            with warnings.catch_warnings():
                warnings.simplefilter('ignore')
                bcf = np.polyfit(sd[idx[:hi]], stress[idx[:hi]], 1)
            bE, bend = bcf[0], hi
            print(f"  ⚠ 强制拟合 E={bE/1000:.2f} GPa（参考）")

        if not math.isnan(bE):
            E_GPa  = bE / 1000.0
            print(f"  -> E={E_GPa:.2f} GPa（{bend} 点）")
            offset = bE * (sd - 0.002) + bcf[1]
            elas_e = sd[idx[bend-1]]
            diff   = stress - offset
            cross  = None
            for i in range(1, n):
                if sd[i] <= elas_e:
                    continue
                if diff[i-1] < 0 <= diff[i]:
                    cross = i; break
            if cross is not None:
                d0, d1 = diff[cross-1], diff[cross]
                t = -d0/(d1-d0) if abs(d1-d0) > 1e-12 else 0.0
                Rp02 = float(stress[cross-1] + t*(stress[cross]-stress[cross-1]))
            else:
                post = sd >= elas_e
                src  = stress[post] if post.any() else stress
                dif  = diff[post]   if post.any() else diff
                Rp02 = float(src[np.argmin(np.abs(dif))])
    except Exception as e:
        print(f"  ⚠ 力学性能计算出错：{e}")

    return {'E': E_GPa, 'Rp02': Rp02, 'Rm': Rm, 'A': A}

def _fmt(v, u=''):
    return f"{v:.3f}{u}" if not math.isnan(v) else 'N/A'

def write_result_xlsx(results, prefix):
    hdrs  = ['载荷', '行程', '引伸计（修正）', '应力', '应变']
    units = ['kN', 'mm', 'mm', 'MPa', '%']
    hfill = PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid')

    mech = {}
    print("\n  各试样力学性能：")
    for name, data in results.items():
        p = calc_mech_props(data)
        mech[name] = p
        print(f"    {name}：Rp0.2={_fmt(p['Rp02'],' MPa')}  Rm={_fmt(p['Rm'],' MPa')}  "
              f"A={_fmt(p['A'],' %')}  E={_fmt(p['E'],' GPa')}")

    def _write(path):
        wb = Workbook()
        wb.remove(wb.active)
        # 汇总 Sheet
        ws0 = wb.create_sheet('力学性能汇总')
        for ci, h in enumerate(['试样','Rp0.2 (MPa)','Rm (MPa)','A (%)','E (GPa)'], 1):
            c = ws0.cell(1, ci, h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.fill = hfill
        for ri, (name, p) in enumerate(mech.items(), 2):
            ws0.cell(ri, 1, name)
            for ci, k in enumerate(['Rp02','Rm','A','E'], 2):
                v = p[k]
                ws0.cell(ri, ci, round(v,3) if not math.isnan(v) else 'N/A')
        for ci in range(1, 6):
            ws0.column_dimensions[get_column_letter(ci)].width = 20
        # 各试样 Sheet
        for name, data in results.items():
            ws = wb.create_sheet(name[:31])
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(1, ci, h)
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center')
                c.fill = hfill
            for ci, u in enumerate(units, 1):
                c = ws.cell(2, ci, u)
                c.font = Font(italic=True, color='555555')
                c.alignment = Alignment(horizontal='center')
            cols = [data['load_kN'], data['stroke_mm'],
                    data['ext_corrected'], data['stress'], data['strain']]
            nrows = max(len(c) for c in cols)
            for ri in range(nrows):
                for ci, col in enumerate(cols, 1):
                    v = col[ri] if ri < len(col) else float('nan')
                    ws.cell(ri+3, ci, None if math.isnan(v) else round(v, 6))
            for ci in range(1, 6):
                ws.column_dimensions[get_column_letter(ci)].width = 16
        wb.save(path)

    out = _write_safe(_write, f"{prefix}_result.xlsx")
    print(f"\n  结果：{out}（{len(results)+1} 个工作表）")

def main():
    print("\n===== 拉伸试验数据处理 v3.0 =====")
    start = input("\n  CSV 文件路径：").strip().strip('"')
    if not os.path.exists(start):
        print("  ✗ 文件不存在"); sys.exit(1)

    folder = os.path.dirname(os.path.abspath(start))
    os.chdir(folder)
    start   = os.path.basename(start)
    stem, _ = os.path.splitext(start)
    m       = re.match(r'^(.+)-(\d+)$', stem)
    prefix  = m.group(1) if m else stem

    print("\n── 查找文件 ──")
    files = find_csv_files(start)
    print("\n── 读取合并 ──")
    rows, specimens = load_and_merge(files)
    print(f"  {len(specimens)} 个试样：{[s[0] for s in specimens]}")
    print("\n── 保存原始数据 ──")
    save_pre_xlsx(rows, specimens, prefix)
    print("\n── 试样参数 ──")
    params = get_all_params(specimens)
    print("\n── 引伸计处理 & 计算 ──")
    results = {}
    for sp_name, col_start in specimens:
        data = extract_specimen(rows, col_start)
        p    = params[sp_name]
        data = process_extensometer(data, p['gauge'], sp_name)
        data = calculate_stress_strain(data, p['area'])
        results[sp_name] = data
    print("\n── 写出结果 ──")
    write_result_xlsx(results, prefix)
    print("\n===== 完成 =====\n")

if __name__ == '__main__':
    main()
