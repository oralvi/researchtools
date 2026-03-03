#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
拉伸试验数据处理脚本 V1.5
依赖：numpy, openpyxl（无 pandas）
"""

import os, re, sys, math, csv, warnings, traceback, time
from datetime import datetime
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

VERSION = "V1.5"
EXT_RANGES = {10: 0.2, 15: 0.3, 25: 0.5}
XLSX_MAX_ROWS = 1048576
PRE_HEADER_ROWS = 2
RESULT_HEADER_ROWS = 2
LOG_LINES = []
RUN_ARTIFACTS = {'pre': None, 'result': None}

def setup_console_encoding():
    # 避免某些 Windows 终端（如 GBK）无法输出特殊符号导致程序崩溃
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(errors='replace')
        except Exception:
            pass

def log(msg):
    print(msg)
    LOG_LINES.append(str(msg))

def write_error_log(target_dir, exc):
    log_path = os.path.join(target_dir, "error.log")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write(f"version: {VERSION}\n")
        f.write(f"time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"error: {repr(exc)}\n\n")
        f.write("traceback:\n")
        f.write(traceback.format_exc())
        f.write("\n\nrecent_log:\n")
        for line in LOG_LINES[-200:]:
            f.write(str(line) + "\n")
    return log_path

def cleanup_artifacts():
    for key in ('pre', 'result'):
        path = RUN_ARTIFACTS.get(key)
        if not path:
            continue
        try:
            if os.path.exists(path):
                os.remove(path)
                log(f"  [INFO] 已删除{key}文件: {path}")
        except OSError:
            log(f"  [WARN] {key}文件删除失败（可能被占用）: {path}")
            log("  [WARN] 请关闭对应Excel文件后手动删除")

def progress_bar(current, total, prefix="", width=20, inline=True):
    total = max(int(total), 1)
    current = min(max(int(current), 0), total)
    filled = int(width * current / total)
    bar = ">" * filled + "." * (width - filled)
    if inline:
        print(f"\r  {prefix}[{bar}] {current}/{total}", end="", flush=True)
    else:
        print(f"  {prefix}[{bar}] {current}/{total}")
    if current == total and inline:
        print()

def _sheet_data_capacity(header_rows):
    return XLSX_MAX_ROWS - header_rows

def _chunk_count(total_rows, capacity):
    return max(1, (total_rows + capacity - 1) // capacity)

def _part_sheet_name(base_name, part_idx, total_parts):
    if total_parts <= 1:
        return base_name[:31]
    suffix = f"_p{part_idx}"
    max_base_len = max(1, 31 - len(suffix))
    return f"{base_name[:max_base_len]}{suffix}"

def _read_raw_csv(filepath):
    for enc in ('gbk', 'gb18030', 'utf-8-sig', 'utf-8', 'latin-1'):
        try:
            with open(filepath, encoding=enc, newline='') as f:
                rows = [
                    [c.strip('"').strip() for c in row]
                    for row in csv.reader(f)
                    if any(c.strip('"').strip() for c in row)
                ]
            log(f"  -> 文件编码：{enc}")
            return rows
        except UnicodeDecodeError:
            continue
    raise IOError(f"无法解码文件：{filepath}")

def detect_specimens(rows):
    row0 = rows[0]
    specimens, col = [], 0
    while col < len(row0):
        name = row0[col].strip()
        if name and name.lower() != 'nan':
            specimens.append((name, col))
            col += 4
        else:
            col += 1
    return specimens

def extract_specimen(rows, col_start):
    data = {'时间': [], '载荷': [], '行程': [], '引伸计': []}
    keys = list(data.keys())
    for row in rows[3:]:
        for ki, key in enumerate(keys):
            ci = col_start + ki
            data[key].append(row[ci] if ci < len(row) else '')
    return data

def find_csv_files(start_file):
    folder = os.path.dirname(os.path.abspath(start_file))
    name, ext = os.path.splitext(os.path.basename(start_file))
    m = re.match(r'^(.+)-(\d+)$', name)
    if not m:
        return [start_file]
    prefix  = m.group(1)
    pattern = re.compile(rf'^{re.escape(prefix)}-(\d+){re.escape(ext)}$', re.I)
    found   = {}
    for fn in os.listdir(folder):
        pm = pattern.match(fn)
        if pm:
            found[int(pm.group(1))] = os.path.join(folder, fn)
    if not found:
        raise FileNotFoundError(f"未找到 {prefix}-*.csv")
    files = [found[k] for k in sorted(found)]
    log(f"\n  发现 {len(files)} 个分段文件：" + ', '.join(os.path.basename(f) for f in files))
    return files

def load_and_merge(files):
    total_files = len(files)
    first_rows = _read_raw_csv(files[0])
    progress_bar(1, total_files, "读取合并 ")
    specimens = detect_specimens(first_rows)
    if total_files > 10 or (len(specimens) > 5 and total_files >= 5):
        log("  [INFO] 正在读取超大数据集，请耐心等待...")
    if total_files == 1:
        return first_rows, specimens

    merged = first_rows[:3] + first_rows[3:]
    for idx, f in enumerate(files[1:], 2):
        rows = _read_raw_csv(f)
        merged.extend(rows[3:])
        progress_bar(idx, total_files, "读取合并 ")
    return merged, specimens

def _safe_path(desired):
    if not os.path.exists(desired):
        return desired
    if input(f"  [WARN] 已存在：{desired}，覆盖？[y/N]: ").strip().lower() == 'y':
        return desired
    base, ext = os.path.splitext(re.sub(r'_\d+$', '', os.path.splitext(desired)[0]))
    ext = os.path.splitext(desired)[1]
    i = 1
    while os.path.exists(f"{base}_{i}{ext}"):
        i += 1
    path = f"{base}_{i}{ext}"
    log(f"  -> 另存为：{path}")
    return path

def _write_safe(fn, desired):
    path = _safe_path(desired)
    while True:
        try:
            fn(path); return path
        except PermissionError:
            base = re.sub(r'_\d+$', '', os.path.splitext(path)[0])
            ext  = os.path.splitext(path)[1]
            i = 1
            while os.path.exists(f"{base}_{i}{ext}"):
                i += 1
            path = f"{base}_{i}{ext}"
            log(f"  [ERR] 文件被占用，另存为：{path}")

def save_pre_xlsx(rows, specimens, prefix):
    def _write(path):
        wb = Workbook()
        wb.remove(wb.active)
        sheet_count = 0
        data_rows = rows[3:]
        total_pre = len(specimens)
        total_units = max(total_pre * max(len(data_rows), 1), 1)
        done_units = 0
        last_percent = 0
        pre_capacity = _sheet_data_capacity(PRE_HEADER_ROWS)
        for idx, (sp_name, col_start) in enumerate(specimens, 1):
            col_labels = [rows[1][col_start+i] if col_start+i < len(rows[1]) else '' for i in range(4)]
            units      = [rows[2][col_start+i] if col_start+i < len(rows[2]) else '' for i in range(4)]
            nrows = len(data_rows)
            part_total = _chunk_count(nrows, pre_capacity)
            if nrows > pre_capacity:
                log(f"  [INFO] pre工作表超行数，试样 {sp_name} 自动拆分为 {part_total} 个分表")

            for part_idx in range(1, part_total + 1):
                ws = wb.create_sheet(_part_sheet_name(sp_name, part_idx, part_total))
                sheet_count += 1
                for ci, lbl in enumerate(col_labels, 1):
                    ws.cell(1, ci, lbl)
                for ci, u in enumerate(units, 1):
                    c = ws.cell(2, ci, u)
                    c.font = Font(italic=True, color='777777')

                start = (part_idx - 1) * pre_capacity
                end = min(start + pre_capacity, nrows)
                for row in data_rows[start:end]:
                    ws.append([
                    row[col_start+ci] if col_start+ci < len(row) else ''
                        for ci in range(4)
                    ])
                    done_units += 1
                    percent = int(done_units * 99 / total_units)
                    if percent > last_percent:
                        last_percent = percent
                        progress_bar(last_percent, 100, "保存pre ")
            if nrows == 0:
                ws = wb.create_sheet(_part_sheet_name(sp_name, 1, 1))
                sheet_count += 1
                for ci, lbl in enumerate(col_labels, 1):
                    ws.cell(1, ci, lbl)
                for ci, u in enumerate(units, 1):
                    c = ws.cell(2, ci, u)
                    c.font = Font(italic=True, color='777777')
                done_units += 1
                percent = int(done_units * 99 / total_units)
                if percent > last_percent:
                    last_percent = percent
                    progress_bar(last_percent, 100, "保存pre ")
        if last_percent < 99:
            progress_bar(99, 100, "保存pre ")
        log("  [INFO] pre数据写入完成，正在保存xlsx文件...")
        wb.save(path)
        progress_bar(100, 100, "保存pre ")
        return sheet_count
    out = _write_safe(_write, f"pre-{prefix}.xlsx")
    log(f"  -> pre 文件：{out}")
    return out

def _calc_area(w, t, cyl):
    if cyl:
        a = w * t * math.pi / 4
        return a, f"圆柱棒 π/4×{w}²={a:.4f} mm²"
    a = w * t
    return a, f"矩形 {w}×{t}={a:.4f} mm²"

def _input_one(name):
    log(f"\n  【{name}】")
    while True:
        try:
            w_txt = input("    宽度/直径 (mm，可留空): ").strip()
            progress_bar(1, 3, "读取参数 ", inline=False)
            tr = input("    厚度 (mm，可留空): ").strip()
            progress_bar(2, 3, "读取参数 ", inline=False)
            g   = float(input("    标距 (mm): "))
            progress_bar(3, 3, "读取参数 ", inline=False)
            w = float(w_txt) if w_txt else float('nan')
            t_raw = float(tr) if tr else float('nan')
            has_w = (not math.isnan(w)) and w > 0
            has_t = (not math.isnan(t_raw)) and t_raw > 0
            if not has_w and not has_t:
                print("    [ERR] 宽度/直径 和 厚度 不能同时为空或<=0")
                continue
            if has_w and has_t:
                cyl = False
                t = t_raw
                if abs(w - t) < 1e-9:
                    cyl = input(f"    宽度=厚度={w}，按圆柱棒处理？[y/N]: ").strip().lower() == 'y'
                    if cyl:
                        t = w
            else:
                cyl = True
                diameter = w if has_w else t_raw
                w = diameter
                t = diameter
                print("    [INFO] 检测到单列直径输入，按圆柱棒处理")
            break
        except ValueError:
            print("    [ERR] 请输入数字")
    a, desc = _calc_area(w, t, cyl)
    if a <= 0:
        print("    [ERR] 截面积<=0，请重新输入")
        return _input_one(name)
    log(f"    -> {desc}，标距={g} mm")
    return {'width': w, 'thickness': t, 'gauge': g, 'area': a}

def _gen_template(specimens, path):
    def _write(target):
        wb = Workbook()
        ws = wb.active
        ws.title = '试样参数'
        hdrs  = ['试样名称', '宽度/直径 (mm)', '厚度 (mm)', '标距 (mm)', '是否圆柱棒(是/否)']
        fills = ['C6EFCE', 'FFEB9C', 'FFEB9C', 'FFEB9C', 'DDEBF7']
        for ci, (h, fc) in enumerate(zip(hdrs, fills), 1):
            c = ws.cell(1, ci, h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.fill = PatternFill(start_color=fc, end_color=fc, fill_type='solid')
        notes = ['<- 勿修改', '<- 矩形填宽，圆柱填直径', '<- 矩形填厚，圆柱留空',
                 '<- 引伸计标距', '<- 填"是"按 π/4·d² 算']
        for ci, n in enumerate(notes, 1):
            ws.cell(2, ci, n).font = Font(italic=True, color='999999', size=9)
        for ri, (sp, _) in enumerate(specimens, 3):
            ws.cell(ri, 1, sp).font = Font(bold=True)
            ws.cell(ri, 5, '否')
        for ci, w in enumerate([20, 18, 14, 14, 22], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        wb.save(target)

    return _write_safe(_write, path)

def _load_template(path, specimens):
    wb  = load_workbook(path, data_only=True)
    ws  = wb.active
    sp_names = [s[0] for s in specimens]
    sp_name_set = set(sp_names)
    params = {}
    loaded = 0
    total = len(sp_names)
    log("\n  参数文件读取中...")
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if name not in sp_name_set or name in params:
            continue
        try:
            g = float(row[3])
        except (TypeError, ValueError):
            raise ValueError(f"第{row_idx}行 {name}: 标距未填写或格式错误")
        w_raw = row[1]
        t_raw = row[2]
        w = float(w_raw) if (w_raw is not None and str(w_raw).strip() != '') else float('nan')
        t = float(t_raw) if (t_raw is not None and str(t_raw).strip() != '') else float('nan')
        has_w = (not math.isnan(w)) and w > 0
        has_t = (not math.isnan(t)) and t > 0
        flag_cyl = str(row[4]).strip().lower() in ('是', 'yes', 'y', '1', 'true')

        if not has_w and not has_t:
            raise ValueError(f"第{row_idx}行 {name}: 宽度/厚度不能同时为空或<=0")

        if has_w and has_t:
            cyl = flag_cyl
            if cyl:
                if abs(w - t) > 1e-9:
                    log(f"  [WARN] {name} 标记为圆柱棒，但宽厚不一致，按宽度/直径列={w} 计算")
                t = w
        else:
            cyl = True
            diameter = w if has_w else t
            w = diameter
            t = diameter
            log(f"  [INFO] {name} 检测到单列直径输入（或另一列为0/空），按圆柱棒处理")

        a, desc = _calc_area(w, t, cyl)
        if a <= 0:
            raise ValueError(f"第{row_idx}行 {name}: 截面积<=0")
        log(f"    {name}：{desc}，标距={g} mm")
        params[name] = {'width': w, 'thickness': t, 'gauge': g, 'area': a}
        loaded += 1
        progress_bar(loaded, total, "读取参数 ", inline=False)
        if loaded == total:
            break
    miss = [s for s in sp_names if s not in params]
    if miss:
        raise ValueError(f"模板中缺少试样参数: {miss}")
    return params

def get_all_params(specimens):
    if len(specimens) == 1:
        return {specimens[0][0]: _input_one(specimens[0][0])}
    tpl = os.path.abspath('specimen_params_template.xlsx')
    print(f"\n-- 多试样参数（{len(specimens)} 个）--")
    print("  [1] 生成模板后填写导入（推荐）")
    print("  [2] 指定已有参数文件")
    print("  [3] 逐个手动输入")
    c = input("  选项 [1/2/3]，默认1：").strip() or '1'
    if c == '3':
        return {n: _input_one(n) for n, _ in specimens}
    if c == '2':
        while True:
            p = input("  参数文件路径：").strip().strip('"')
            if not p:
                print("  [ERR] 未填写内容，请输入参数文件路径")
                continue
            if not os.path.exists(p):
                print("  [ERR] 文件不存在")
                continue
            try:
                return _load_template(p, specimens)
            except ValueError as e:
                print(f"  [ERR] {e}")
                cmd = input("  [INFO] 按回车重新选择文件，输入 q 退出参数输入: ").strip().lower()
                if cmd == 'q':
                    raise RuntimeError("用户取消参数输入")
    else:
        tpl = _gen_template(specimens, tpl)
        while True:
            print(f"\n  模板：{tpl}")
            print("  填写宽度/直径、厚度、标距、是否圆柱棒后保存，按回车继续...")
            input()
            try:
                params = _load_template(tpl, specimens)
                try:
                    if os.path.exists(tpl):
                        os.remove(tpl)
                        log(f"  [INFO] 已删除参数模板：{tpl}")
                except OSError:
                    log(f"  [WARN] 参数模板删除失败（可能被Excel占用）：{tpl}")
                    log("  [WARN] 请关闭对应Excel文件后手动删除")
                return params
            except ValueError as e:
                print(f"  [ERR] {e}")
                cmd = input("  [INFO] 检测到未填写内容或格式错误。按回车继续修改模板，输入 q 退出: ").strip().lower()
                if cmd == 'q':
                    raise RuntimeError("用户取消参数输入")

def _pf(s):
    s = str(s).strip()
    if s in ('-.----', '', 'nan', '-'):
        return float('nan')
    try:
        return float(s)
    except ValueError:
        return float('nan')

def process_extensometer(data, gauge, sp_name):
    print(f"\n  [引伸计] {sp_name}")
    raw_ext = data['引伸计']
    ext_num = [_pf(v) for v in raw_ext]

    first_valid = next((v for v in ext_num if not math.isnan(v)), float('nan'))
    if math.isnan(first_valid) or first_valid == 0.0:
        if not any(not math.isnan(v) and v != 0.0 for v in ext_num):
            if input("  [WARN] 引伸计无效，继续？[y/N]: ").strip().lower() != 'y':
                raise RuntimeError(f"{sp_name} 引伸计无效且用户取消处理")
            data['ext_corrected'] = [float('nan')] * len(raw_ext)
            data['ext_mm'] = 10.0
            return data

    corrected = list(ext_num)
    cum_off, prev, slips = 0.0, None, 0
    for i, v in enumerate(ext_num):
        if math.isnan(v):
            continue
        if prev is not None and (v - prev) < -0.05:
            cum_off += prev - v
            slips   += 1
            print(f"  [WARN] 滑脱 第{i+1}行：{prev:.4f}->{v:.4f} 偏移={cum_off:.4f}")
        corrected[i] = v + cum_off
        prev = v
    if slips:
        print(f"  共 {slips} 次滑脱，已修正")

    dash_idx = next((i for i, v in enumerate(raw_ext) if str(v).strip() == '-.----'), None)
    ext_mm   = None

    if dash_idx is not None and dash_idx > 0:
        last_i   = max(i for i in range(dash_idx) if not math.isnan(corrected[i]))
        last_val = corrected[last_i]
        ext_mm   = min(EXT_RANGES, key=lambda mm: abs(last_val - EXT_RANGES[mm]))
        print(f"  -> {ext_mm}mm 引伸计（末值={last_val:.4f}，满量程≈{EXT_RANGES[ext_mm]}）")
        try:
            last_stroke = _pf(data['行程'][last_i])
        except (IndexError, ValueError):
            last_stroke = 0.0
        print(f"  -> 行程替代：(行程-{last_stroke:.4f})/{gauge}×{ext_mm}+{last_val:.4f}")
        for i in range(dash_idx, len(corrected)):
            if math.isnan(corrected[i]):
                s = _pf(data['行程'][i])
                if not math.isnan(s):
                    corrected[i] = (s - last_stroke) / gauge * ext_mm + last_val

    data['ext_corrected'] = corrected
    data['ext_mm']        = float(ext_mm) if ext_mm else 10.0
    return data

def calculate_stress_strain(data, area):
    ext_mm = data['ext_mm']
    load_arr = np.array([_pf(v) for v in data['载荷']], dtype=float)
    stroke_arr = np.array([_pf(v) for v in data['行程']], dtype=float)
    ext_arr = np.array(data['ext_corrected'], dtype=float)

    stress_arr = load_arr / area * 1000.0
    strain_arr = ext_arr / ext_mm * 100.0

    data['load_kN'] = load_arr.tolist()
    data['stroke_mm'] = stroke_arr.tolist()
    data['stress'] = stress_arr.tolist()
    data['strain'] = strain_arr.tolist()
    # 释放原始字符串列，降低内存占用
    data.pop('时间', None)
    data.pop('载荷', None)
    data.pop('行程', None)
    data.pop('引伸计', None)
    data.pop('ext_mm', None)
    return data

def calc_mech_props(data):
    nan = float('nan')
    stress = np.array([v for v in data['stress'] if not math.isnan(v)])
    strain = np.array([v for v in data['strain'] if not math.isnan(v)])
    n = min(len(stress), len(strain))
    stress, strain = stress[:n], strain[:n]
    if n < 10:
        return {'E': nan, 'Rp02': nan, 'Rm': nan, 'A': nan}

    Rm = float(np.max(stress))
    A  = float(np.max(strain))
    E_GPa, Rp02 = nan, nan
    fit_tag = '仅供参考'
    fit_points = 0
    fit_r2 = nan
    fit_strain_max_pct = nan

    def _fit_r2(x, y):
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            cf = np.polyfit(x, y, 1)
        fit = np.polyval(cf, x)
        sst = np.sum((y - np.mean(y)) ** 2)
        sse = np.sum((y - fit) ** 2)
        r2 = 1.0 - sse / sst if sst > 1e-12 else 1.0
        return cf, r2

    try:
        sd = strain / 100.0
        min_pts = 8
        # 约束在较早应变区间，避免把塑性段拉进弹性拟合
        for strain_cap in (0.010, 0.020):
            idx = np.where(
                (stress >= 0.05 * Rm) &
                (stress <= 0.70 * Rm) &
                (sd <= strain_cap)
            )[0]
            if len(idx) < min_pts:
                continue

            sa, sta = sd[idx], stress[idx]
            best = None
            for end in range(min_pts, len(sa) + 1):
                cf, r2 = _fit_r2(sa[:end], sta[:end])
                score = (r2, -end)  # r2优先；点数少优先，避免过长段
                if best is None or score > best['score']:
                    best = {'cf': cf, 'r2': r2, 'end': end, 'idx': idx, 'score': score}

            if best is not None:
                bcf = best['cf']
                bE = bcf[0]
                bend = best['end']
                fit_end_idx = best['idx'][bend - 1]
                fit_end_strain = sd[fit_end_idx]
                fit_points = bend
                fit_r2 = best['r2']
                fit_strain_max_pct = fit_end_strain * 100.0
                E_GPa = bE / 1000.0
                fit_tag = '精确拟合' if best['r2'] >= 0.9995 else '仅供参考'
                print(f"  -> E={E_GPa:.2f} GPa（{bend} 点，R2={best['r2']:.6f}，{fit_tag}）")

                offset = bE * (sd - 0.002) + bcf[1]
                diff = stress - offset
                cross = None
                for i in range(1, n):
                    if sd[i] <= fit_end_strain:
                        continue
                    if diff[i - 1] < 0 <= diff[i]:
                        cross = i
                        break
                if cross is not None:
                    d0, d1 = diff[cross - 1], diff[cross]
                    t = -d0 / (d1 - d0) if abs(d1 - d0) > 1e-12 else 0.0
                    Rp02 = float(stress[cross - 1] + t * (stress[cross] - stress[cross - 1]))
                else:
                    post = sd >= fit_end_strain
                    src = stress[post] if post.any() else stress
                    dif = diff[post] if post.any() else diff
                    Rp02 = float(src[np.argmin(np.abs(dif))])
                break
    except Exception as e:
        print(f"  [WARN] 力学性能计算出错：{e}")

    return {
        'E': E_GPa,
        'Rp02': Rp02,
        'Rm': Rm,
        'A': A,
        'fit_tag': fit_tag,
        'fit_points': fit_points,
        'fit_r2': fit_r2,
        'fit_strain_max_pct': fit_strain_max_pct,
    }

def _fmt(v, u=''):
    return f"{v:.3f}{u}" if not math.isnan(v) else 'N/A'

def write_result_xlsx(results, prefix):
    hdrs  = ['载荷', '行程', '引伸计（修正）', '应力', '应变']
    units = ['kN', 'mm', 'mm', 'MPa', '%']
    hfill = PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid')

    mech = {}
    log("\n  各试样力学性能：")
    for name, data in results.items():
        p = calc_mech_props(data)
        mech[name] = p
        log(f"    {name}：Rp0.2={_fmt(p['Rp02'],' MPa')}  Rm={_fmt(p['Rm'],' MPa')}  "
            f"A={_fmt(p['A'],' %')}  E={_fmt(p['E'],' GPa')}（{p['fit_tag']}）")
        if p['fit_points'] > 0 and not math.isnan(p['fit_r2']) and not math.isnan(p['fit_strain_max_pct']):
            log(
                f"      拟合区间点: {int(p['fit_points'])}, "
                f"R2={p['fit_r2']:.6f}, 应变上限={p['fit_strain_max_pct']:.4f}%"
            )
        else:
            log("      拟合区间点: N/A（未获得有效弹性拟合区间）")

    def _write(path):
        wb = Workbook()
        wb.remove(wb.active)
        # 首页日志 Sheet（必须第一个）
        ws_log = wb.create_sheet('运行日志')
        ws_log.cell(1, 1, '程序版本').font = Font(bold=True)
        ws_log.cell(1, 2, VERSION)
        ws_log.cell(2, 1, '生成时间').font = Font(bold=True)
        ws_log.cell(2, 2, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ws_log.cell(3, 1, '日志内容').font = Font(bold=True)
        for ri, line in enumerate(LOG_LINES, 4):
            ws_log.cell(ri, 1, line)
        ws_log.column_dimensions[get_column_letter(1)].width = 120
        ws_log.column_dimensions[get_column_letter(2)].width = 24

        # 汇总 Sheet
        ws0 = wb.create_sheet('力学性能汇总')
        for ci, h in enumerate(['试样','Rp0.2 (MPa)','Rm (MPa)','A (%)','E (GPa)'], 1):
            c = ws0.cell(1, ci, h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.fill = hfill
        for ri, (name, p) in enumerate(mech.items(), 2):
            ws0.cell(ri, 1, name)
            for ci, k in enumerate(['Rp02','Rm','A','E'], 2):
                v = p[k]
                ws0.cell(ri, ci, round(v,3) if not math.isnan(v) else 'N/A')
        for ci in range(1, 6):
            ws0.column_dimensions[get_column_letter(ci)].width = 20
        # 各试样 Sheet（按数据写入量做百分比进度）
        save_work = []
        total_units = 0
        result_capacity = _sheet_data_capacity(RESULT_HEADER_ROWS)
        for name, data in results.items():
            cols = (
                data['load_kN'],
                data['stroke_mm'],
                data['ext_corrected'],
                data['stress'],
                data['strain'],
            )
            nrows = max((len(c) for c in cols), default=0)
            save_work.append((name, cols, nrows))
            total_units += max(nrows, 1)
        total_units = max(total_units, 1)
        done_units = 0
        last_percent = 0

        for name, cols, nrows in save_work:
            part_total = _chunk_count(nrows, result_capacity)
            if nrows > result_capacity:
                log(f"  [INFO] result工作表超行数，试样 {name} 自动拆分为 {part_total} 个分表")

            for part_idx in range(1, part_total + 1):
                ws = wb.create_sheet(_part_sheet_name(name, part_idx, part_total))
                for ci, h in enumerate(hdrs, 1):
                    c = ws.cell(1, ci, h)
                    c.font = Font(bold=True)
                    c.alignment = Alignment(horizontal='center')
                    c.fill = hfill
                for ci, u in enumerate(units, 1):
                    c = ws.cell(2, ci, u)
                    c.font = Font(italic=True, color='555555')
                    c.alignment = Alignment(horizontal='center')

                start = (part_idx - 1) * result_capacity
                end = min(start + result_capacity, nrows)
                for ri in range(start, end):
                    row_vals = []
                    for col in cols:
                        v = col[ri] if ri < len(col) else float('nan')
                        row_vals.append(None if math.isnan(v) else round(v, 6))
                    ws.append(row_vals)
                    done_units += 1
                    percent = int(done_units * 99 / total_units)
                    if percent > last_percent:
                        last_percent = percent
                        progress_bar(last_percent, 100, "保存进度 ")
                for ci in range(1, 6):
                    ws.column_dimensions[get_column_letter(ci)].width = 16

            if nrows == 0:
                ws = wb.create_sheet(_part_sheet_name(name, 1, 1))
                for ci, h in enumerate(hdrs, 1):
                    c = ws.cell(1, ci, h)
                    c.font = Font(bold=True)
                    c.alignment = Alignment(horizontal='center')
                    c.fill = hfill
                for ci, u in enumerate(units, 1):
                    c = ws.cell(2, ci, u)
                    c.font = Font(italic=True, color='555555')
                    c.alignment = Alignment(horizontal='center')
                for ci in range(1, 6):
                    ws.column_dimensions[get_column_letter(ci)].width = 16
                done_units += 1
                percent = int(done_units * 99 / total_units)
                if percent > last_percent:
                    last_percent = percent
                    progress_bar(last_percent, 100, "保存进度 ")
        if last_percent < 99:
            progress_bar(99, 100, "保存进度 ")
        log("  [INFO] result数据写入完成，正在保存xlsx文件...")
        wb.save(path)
        progress_bar(100, 100, "保存进度 ")

    out = _write_safe(_write, f"{prefix}_result.xlsx")
    total_sheets = 0
    try:
        wb_check = load_workbook(out, read_only=True)
        total_sheets = len(wb_check.sheetnames)
        wb_check.close()
    except Exception:
        total_sheets = len(results) + 2
    log(f"\n  结果：{out}（{total_sheets} 个工作表）")
    return out

def _process_one(start):
    folder = os.path.dirname(os.path.abspath(start))
    os.chdir(folder)
    start   = os.path.basename(start)
    stem, _ = os.path.splitext(start)
    m       = re.match(r'^(.+)-(\d+)$', stem)
    prefix  = m.group(1) if m else stem

    log("\n-- 查找文件 --")
    files = find_csv_files(start)
    log("\n-- 读取合并 --")
    rows, specimens = load_and_merge(files)
    log(f"  {len(specimens)} 个试样：{[s[0] for s in specimens]}")
    log("\n-- 保存原始数据 --")
    pre_path = save_pre_xlsx(rows, specimens, prefix)
    RUN_ARTIFACTS['pre'] = pre_path
    log("\n-- 试样参数 --")
    params = get_all_params(specimens)
    log("\n-- 引伸计处理 & 计算 --")
    results = {}
    for sp_name, col_start in specimens:
        data = extract_specimen(rows, col_start)
        p    = params[sp_name]
        data = process_extensometer(data, p['gauge'], sp_name)
        data = calculate_stress_strain(data, p['area'])
        results[sp_name] = {
            'load_kN': data['load_kN'],
            'stroke_mm': data['stroke_mm'],
            'ext_corrected': data['ext_corrected'],
            'stress': data['stress'],
            'strain': data['strain'],
        }
    log("\n-- 写出结果 --")
    result_path = write_result_xlsx(results, prefix)
    RUN_ARTIFACTS['result'] = result_path

    keep_pre = input("\n  是否保留 pre xlsx？[y/N]: ").strip().lower() == 'y'
    if keep_pre:
        log(f"  -> 保留 pre 文件：{pre_path}")
    else:
        if pre_path and os.path.exists(pre_path):
            try:
                os.remove(pre_path)
                log(f"  -> 已删除 pre 文件：{pre_path}")
            except OSError as e:
                log(f"  [WARN] pre 文件删除失败：{e}")
        else:
            log("  -> 未找到 pre 文件，无需删除")

def main():
    log(f"\n===== 拉伸试验数据处理 {VERSION} =====")
    while True:
        start = input("\n  CSV 文件路径：").strip().strip('"')
        if not start:
            log("  [ERR] 请输入文件路径")
            continue
        if not os.path.exists(start):
            log("  [ERR] 文件不存在")
            continue
        RUN_ARTIFACTS['pre'] = None
        RUN_ARTIFACTS['result'] = None
        try:
            _process_one(start)
        except Exception as e:
            target_dir = os.path.dirname(os.path.abspath(start))
            log_path = write_error_log(target_dir, e)
            log(f"  [ERR] 本次处理失败，详情见错误日志: {log_path}")
            cleanup_artifacts()
            log("  [INFO] 程序将在 7 秒后自动退出")
            time.sleep(7)
            break

        again = input("\n  还有需要处理的文件吗？[y/N]: ").strip().lower() == 'y'
        if not again:
            break

    log("\n===== 完成 =====\n")

if __name__ == '__main__':
    setup_console_encoding()
    main()


