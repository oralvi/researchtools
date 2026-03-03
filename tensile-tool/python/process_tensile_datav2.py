#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
拉伸试验数据处理脚本 v2.0
支持：
  - 单试样 / 多试样（单 CSV 内多组数据）
  - 多分段文件合并（20231120-1.csv ... 20231120-N.csv）
  - 引伸计滑脱检测与修正
  - 引伸计量程识别（10/15/25mm）及行程替代
  - 应力（MPa）/ 应变（%）计算
  - 多试样分 Sheet 输出 pre-*.xlsx 和 *_result.xlsx
"""

import os
import re
import sys
import math
import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────────────────
# 工具：安全写出 xlsx（处理覆盖提示 & 文件占用）
# ──────────────────────────────────────────────────────────
def _safe_output_path(desired: str) -> str:
    """
    若文件已存在，询问是否覆盖。
    不覆盖则自动生成 name_1.xlsx / name_2.xlsx ...
    """
    if not os.path.exists(desired):
        return desired
    ans = input(f"  ⚠ 文件已存在：{desired}，是否覆盖？[y/N]: ").strip().lower()
    if ans == 'y':
        return desired
    # 自动生成新文件名
    base, ext = os.path.splitext(desired)
    i = 1
    while True:
        candidate = f"{base}_{i}{ext}"
        if not os.path.exists(candidate):
            print(f"  → 将另存为：{candidate}")
            return candidate
        i += 1


def _write_xlsx_safe(write_func, desired_name: str, *args, **kwargs):
    """
    调用 write_func(out_name, *args, **kwargs) 写文件，
    自动处理：① 覆盖确认；② 文件被占用时提示并另存。
    write_func 签名：write_func(out_name: str, ...) -> None
    """
    out_name = _safe_output_path(desired_name)
    while True:
        try:
            write_func(out_name, *args, **kwargs)
            return out_name
        except PermissionError:
            print(f"  ✗ 无法写入 {out_name}，文件可能已在 Excel 中打开，自动另存为新文件名。")
            base, ext = os.path.splitext(out_name)
            # 去掉已有的 _数字 后缀再重新编号，避免 _1_1_1...
            base = re.sub(r'_\d+$', '', base)
            i = 1
            while True:
                out_name = f"{base}_{i}{ext}"
                if not os.path.exists(out_name):
                    print(f"  → 另存为：{out_name}")
                    break
                i += 1


# ──────────────────────────────────────────────────────────
# 工具：读取原始 CSV（GBK 兼容）
# ──────────────────────────────────────────────────────────
def _strip_df(df):
    """去除所有字符串单元格的引号和首尾空格，兼容 pandas 新旧版本。"""
    def _clean(x):
        return x.strip('"').strip() if isinstance(x, str) else x
    try:
        return df.applymap(_clean)   # pandas < 2.1
    except AttributeError:
        return df.map(_clean)        # pandas >= 2.1


def _read_raw_csv(filepath: str) -> pd.DataFrame:
    for enc in ['gbk', 'gb18030', 'gb2312', 'utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
        try:
            df = pd.read_csv(filepath, header=None, encoding=enc, dtype=str,
                             on_bad_lines='skip')
            df = _strip_df(df)
            print(f"  → 文件编码识别为：{enc}")
            return df
        except UnicodeDecodeError:
            continue   # 只跳过真正的解码错误
    raise IOError(f"无法解码文件：{filepath}")


# ──────────────────────────────────────────────────────────
# 1. 检测试样数量及名称
#    第1行：每4列的第0列是试样名称（其余3列为空）
# ──────────────────────────────────────────────────────────
def detect_specimens(raw: pd.DataFrame) -> list:
    row0 = raw.iloc[0].tolist()
    specimens = []
    col = 0
    while col < len(row0):
        name = str(row0[col]).strip()
        if name and name not in ('', 'nan'):
            specimens.append((name, col))
            col += 4
        else:
            col += 1
    return specimens


# ──────────────────────────────────────────────────────────
# 2. 拆出单个试样的 4 列数据（跳过前3行 header）
# ──────────────────────────────────────────────────────────
def extract_specimen_df(raw: pd.DataFrame, col_start: int) -> pd.DataFrame:
    sub = raw.iloc[:, col_start:col_start + 4].copy()
    sub.columns = ['时间', '载荷', '行程', '引伸计']
    data = sub.iloc[3:].reset_index(drop=True)
    return data


# ──────────────────────────────────────────────────────────
# 3. 文件发现（单文件 / 分段文件）
# ──────────────────────────────────────────────────────────
def find_csv_files(start_file: str) -> list:
    folder = os.path.dirname(os.path.abspath(start_file))
    basename = os.path.basename(start_file)
    name, ext = os.path.splitext(basename)

    m = re.match(r'^(.+)-(\d+)$', name)
    if not m:
        if os.path.exists(start_file):
            return [start_file]
        raise FileNotFoundError(f"文件不存在：{start_file}")

    prefix = m.group(1)
    pattern = re.compile(rf'^{re.escape(prefix)}-(\d+){re.escape(ext)}$', re.IGNORECASE)
    found = {}
    for fn in os.listdir(folder):
        pm = pattern.match(fn)
        if pm:
            found[int(pm.group(1))] = os.path.join(folder, fn)

    if not found:
        raise FileNotFoundError(f"未找到 {prefix}-*.csv")

    sorted_files = [found[k] for k in sorted(found.keys())]
    print(f"\n  发现 {len(sorted_files)} 个分段文件：")
    for f in sorted_files:
        print(f"    {os.path.basename(f)}")
    return sorted_files


# ──────────────────────────────────────────────────────────
# 4. 读取 & 合并所有分段文件
# ──────────────────────────────────────────────────────────
def load_and_merge_files(files: list) -> tuple:
    raws = [_read_raw_csv(f) for f in files]
    specimens = detect_specimens(raws[0])

    if len(files) == 1:
        return raws[0], specimens

    header_rows = raws[0].iloc[:3].copy()
    data_parts = [raw.iloc[3:] for raw in raws]
    merged_data = pd.concat(data_parts, ignore_index=True)
    merged_raw = pd.concat([header_rows, merged_data], ignore_index=True)
    return merged_raw, specimens


# ──────────────────────────────────────────────────────────
# 5. 保存 pre-*.xlsx（每试样一个 Sheet，含单位行）
# ──────────────────────────────────────────────────────────
def save_pre_xlsx(raw: pd.DataFrame, specimens: list, prefix: str) -> str:
    desired = f"pre-{prefix}.xlsx"

    def _do_write(out_name):
        with pd.ExcelWriter(out_name, engine='openpyxl') as writer:
            for sp_name, col_start in specimens:
                col_labels = raw.iloc[1, col_start:col_start + 4].tolist()
                units = raw.iloc[2, col_start:col_start + 4].tolist()
                df = extract_specimen_df(raw, col_start)
                df.columns = col_labels
                df.to_excel(writer, sheet_name=sp_name[:31], index=False)
                ws = writer.sheets[sp_name[:31]]
                ws.insert_rows(2)
                for i, u in enumerate(units, start=1):
                    cell = ws.cell(row=2, column=i, value=u)
                    cell.font = Font(italic=True, color='777777')

    out_name = _write_xlsx_safe(_do_write, desired)
    print(f"  → 已保存 pre 文件：{out_name}（{len(specimens)} 个工作表）")
    return out_name


# ──────────────────────────────────────────────────────────
# 6. 试样参数输入（单试样=命令行交互；多试样=xlsx模板）
# ──────────────────────────────────────────────────────────
def _calc_area(width: float, thickness: float, is_cylinder: bool) -> tuple:
    """计算截面积，返回 (area, 描述文字)。"""
    if is_cylinder:
        area = width * thickness * math.pi / 4
        return area, f"圆柱棒 S=π/4×{width}²={area:.4f} mm²"
    else:
        area = width * thickness
        return area, f"矩形截面 S={width}×{thickness}={area:.4f} mm²"


def _input_single_specimen(sp_name: str) -> dict:
    """命令行逐个输入单个试样参数。"""
    print(f"\n  【试样：{sp_name}】")
    while True:
        try:
            width = float(input(f"    宽度/直径 (mm): "))
            t_raw = input(f"    厚度 (mm)（圆柱棒直接回车）: ").strip()
            if t_raw == '':
                thickness, is_cylinder = width, True
            else:
                thickness, is_cylinder = float(t_raw), False
            gauge = float(input(f"    标距 (mm): "))
            break
        except ValueError:
            print("    ✗ 请输入有效数字。")

    if not is_cylinder and abs(width - thickness) < 1e-9:
        ans = input(f"    宽度=厚度={width}mm，是否为圆柱棒？[y/N]: ").strip().lower()
        is_cylinder = (ans == 'y')

    area, desc = _calc_area(width, thickness, is_cylinder)
    print(f"    → {desc}")
    return {'width': width, 'thickness': thickness, 'gauge': gauge,
            'area': area, 'is_cylinder': is_cylinder}


def _generate_template(specimens: list, template_path: str):
    """生成参数模板 xlsx，含示例数据和说明。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = '试样参数'

    # 标题行
    headers = ['试样名称', '宽度/直径 (mm)', '厚度 (mm)', '标距 (mm)',
               '是否圆柱棒 (是/否)']
    fills = ['C6EFCE', 'FFEB9C', 'FFEB9C', 'FFEB9C', 'DDEBF7']
    for ci, (h, fc) in enumerate(zip(headers, fills), start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
        c.fill = PatternFill(start_color=fc, end_color=fc, fill_type='solid')

    # 说明行
    notes = ['← 自动填入，请勿修改',
             '← 矩形填宽度，圆柱填直径',
             '← 矩形填厚度，圆柱与宽度相同或留空',
             '← 引伸计标距长度',
             '← 填"是"则按 π/4×d² 计算截面积']
    for ci, note in enumerate(notes, start=1):
        c = ws.cell(row=2, column=ci, value=note)
        c.font = Font(italic=True, color='888888', size=9)

    # 各试样行（预填名称，其余留空等用户填写）
    for ri, (sp_name, _) in enumerate(specimens, start=3):
        ws.cell(row=ri, column=1, value=sp_name).font = Font(bold=True)
        ws.cell(row=ri, column=5, value='否')   # 默认非圆柱棒

    # 列宽
    widths = [20, 18, 14, 14, 22]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(template_path)


def _load_params_from_xlsx(xlsx_path: str, specimens: list) -> dict:
    """从填好的模板 xlsx 读取各试样参数。"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    # 第3行起是数据（行1=标题，行2=说明）
    sp_names = [s[0] for s in specimens]
    params = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if name not in sp_names:
            continue
        try:
            width     = float(row[1])
            t_raw     = row[2]
            gauge     = float(row[3])
            cyl_flag  = str(row[4]).strip() if row[4] is not None else '否'
        except (TypeError, ValueError) as e:
            print(f"  ✗ 试样 {name} 参数读取失败：{e}，请检查模板文件。")
            sys.exit(1)

        if t_raw is None or str(t_raw).strip() == '':
            thickness, is_cylinder = width, True
        else:
            thickness = float(t_raw)
            is_cylinder = cyl_flag in ('是', 'yes', 'Yes', 'YES', 'y', 'Y', '1', 'true', 'True')
            if not is_cylinder and abs(width - thickness) < 1e-9:
                is_cylinder = True   # 宽厚相同视为圆柱棒

        area, desc = _calc_area(width, thickness, is_cylinder)
        print(f"    {name}：{desc}，标距={gauge} mm")
        params[name] = {'width': width, 'thickness': thickness,
                        'gauge': gauge, 'area': area, 'is_cylinder': is_cylinder}

    missing = [s for s in sp_names if s not in params]
    if missing:
        print(f"  ✗ 以下试样在模板中未找到：{missing}")
        sys.exit(1)
    return params


def get_all_specimen_params(specimens: list) -> dict:
    """
    单试样：命令行逐个输入。
    多试样：生成/读取 xlsx 模板。
    """
    if len(specimens) == 1:
        # 单试样直接命令行输入
        print("\n─── 请输入试样尺寸参数（单位：mm）───")
        sp_name = specimens[0][0]
        return {sp_name: _input_single_specimen(sp_name)}

    # 多试样：模板流程
    template_path = os.path.abspath('specimen_params_template.xlsx')
    print(f"\n─── 多试样参数输入 ───")
    print(f"  检测到 {len(specimens)} 个试样，推荐使用 xlsx 模板批量填写。")
    print(f"\n  请选择输入方式：")
    print(f"    [1] 自动生成模板，填写后导入（推荐）")
    print(f"    [2] 指定已有的参数 xlsx 文件路径")
    print(f"    [3] 逐个手动输入")
    choice = input("  请输入选项 [1/2/3]，默认1：").strip() or '1'

    if choice == '3':
        print("\n─── 逐个输入试样参数（单位：mm）───")
        return {sp_name: _input_single_specimen(sp_name) for sp_name, _ in specimens}

    if choice == '2':
        xlsx_path = input("  请输入参数文件路径：").strip().strip('"')
        if not os.path.exists(xlsx_path):
            print(f"  ✗ 文件不存在：{xlsx_path}")
            sys.exit(1)
    else:
        # 选项1：生成模板
        _generate_template(specimens, template_path)
        print(f"\n  ✓ 已生成模板文件：{template_path}")
        print(f"\n  ┌─────────────────────────────────────────────────┐")
        print(f"  │  请打开模板，按列填写各试样的宽度/直径、厚度、   │")
        print(f"  │  标距，以及是否圆柱棒（是/否），然后保存。       │")
        print(f"  │                                                   │")
        print(f"  │  注意：试样名称列请勿修改！                      │")
        print(f"  └─────────────────────────────────────────────────┘")
        input("\n  填写完毕后按回车继续...")
        xlsx_path = template_path

    print(f"\n  正在读取参数文件：{xlsx_path}")
    return _load_params_from_xlsx(xlsx_path, specimens)


# ──────────────────────────────────────────────────────────
# 7. 引伸计检测与修正（单试样）
# ──────────────────────────────────────────────────────────
EXT_RANGES = {10: 0.2, 15: 0.3, 25: 0.5}

def process_extensometer(df: pd.DataFrame, gauge: float, sp_name: str) -> pd.DataFrame:
    print(f"\n  [引伸计] 试样 {sp_name}")
    ext_col = df['引伸计'].copy()

    def parse_ext(val):
        s = str(val).strip()
        if s in ('-.----', '', 'nan'):
            return np.nan
        try:
            return float(s)
        except:
            return np.nan

    ext_num = ext_col.apply(parse_ext)

    # 检查首值（0 或 NaN 表示引伸计无效）
    first_valid = ext_num.dropna().iloc[0] if not ext_num.dropna().empty else np.nan
    if np.isnan(first_valid) or first_valid == 0.0:
        # 首值恰好为 0 还需再看：若整列都是 0，才真正无效
        non_zero = (ext_num != 0) & ext_num.notna()
        if not non_zero.any():
            print(f"  ⚠ 引伸计不存在或有问题（首值={ext_col.iloc[0]}），是否继续？[y/N]: ", end='')
            if input().strip().lower() != 'y':
                print("  已终止。")
                sys.exit(0)
            df['引伸计_修正'] = np.nan
            df['引伸计类型_mm'] = np.nan
            return df

    # ── 滑脱检测与修正（只修正有数值的行，-.---- 不动）──
    ext_corrected = ext_num.copy()   # NaN 表示原始 -.----，保持不动
    cum_off = 0.0
    prev_raw = None
    slip_count = 0

    for i in range(len(ext_num)):
        v = ext_num.iloc[i]
        if np.isnan(v):
            continue   # -.---- 行跳过，不修改
        if prev_raw is not None:
            delta = v - prev_raw
            if delta < -0.05:
                cum_off += (prev_raw - v)
                slip_count += 1
                print(f"  ⚠ 引伸计滑脱：第 {i+1} 行，"
                      f"前值={prev_raw:.4f} → 后值={v:.4f}，累计偏移={cum_off:.4f}")
                print("  → 数据可能不正确，将对后续行进行偏移修正。")
        ext_corrected.iloc[i] = v + cum_off   # 有数值的行：加偏移（无滑脱时 cum_off=0，值不变）
        prev_raw = v

    if slip_count:
        print(f"  共检测到 {slip_count} 次滑脱，已修正。")

    # ── 查找第一个 "-.----"，识别引伸计量程，替代后续 -.---- 行 ──
    dashline_idx = None
    for i in range(len(ext_col)):
        if str(ext_col.iloc[i]).strip() == '-.----':
            dashline_idx = i
            break

    ext_type_mm = None
    if dashline_idx is not None and dashline_idx > 0:
        # 取 -.---- 前最后一个有效值（已含滑脱修正）
        valid_before = ext_corrected.iloc[:dashline_idx].dropna()
        if not valid_before.empty:
            last_valid_val = valid_before.iloc[-1]          # 原始引伸计无量纲值，如 0.5
            last_valid_idx = valid_before.index[-1]

            # 匹配引伸计量程
            min_diff = float('inf')
            for mm, norm_val in EXT_RANGES.items():
                diff = abs(last_valid_val - norm_val)
                if diff < min_diff:
                    min_diff = diff
                    ext_type_mm = mm

            print(f"  → 识别到存在 {ext_type_mm}mm 引伸计"
                  f"（末值={last_valid_val:.4f}，满量程≈{EXT_RANGES[ext_type_mm]}）")

            try:
                last_stroke = float(df['行程'].iloc[last_valid_idx])
            except:
                last_stroke = 0.0

            # 只替换 -.---- 的行，原始有数值的行绝对不动
            print(f"  → 从第 {dashline_idx+1} 行起替代 -.---- 行")
            print(f"     公式：(行程 - {last_stroke:.4f}) / {gauge:.1f} × {ext_type_mm} + {last_valid_val:.4f}")
            for i in range(dashline_idx, len(ext_corrected)):
                if np.isnan(ext_corrected.iloc[i]):   # 只处理 -.---- 对应的 NaN 行
                    try:
                        stroke_i = float(df['行程'].iloc[i])
                        ext_corrected.iloc[i] = (stroke_i - last_stroke) / gauge * ext_type_mm + last_valid_val
                    except:
                        pass   # 保持 NaN

    _ext_mm = float(ext_type_mm) if ext_type_mm else 10.0
    df['引伸计_修正'] = ext_corrected
    df['引伸计类型_mm'] = _ext_mm
    return df


# ──────────────────────────────────────────────────────────
# 8. 计算应力与应变（单试样）
# ──────────────────────────────────────────────────────────
def calculate_stress_strain(df: pd.DataFrame, area: float, gauge: float) -> pd.DataFrame:
    ext_mm = float(df['引伸计类型_mm'].iloc[0]) if pd.notna(df['引伸计类型_mm'].iloc[0]) else 0
    df['载荷_kN'] = pd.to_numeric(df['载荷'], errors='coerce')
    df['行程_mm'] = pd.to_numeric(df['行程'], errors='coerce')
    df['应力_MPa'] = df['载荷_kN'] / area * 1000
    if ext_mm > 0:
        # 引伸计列为原始无量纲值（仪器直接输出），应变(%) = 引伸计值 / 引伸计标距(mm) × 100
        df['应变_%'] = df['引伸计_修正'] / ext_mm * 100
    else:
        df['应变_%'] = np.nan
    return df


# ──────────────────────────────────────────────────────────
# 9. 力学性能参数计算（E、Rp0.2、Rm、A）
# ──────────────────────────────────────────────────────────
def calc_mechanical_properties(df: pd.DataFrame) -> dict:
    """
    计算：E (GPa)、Rp0.2 (MPa)、Rm (MPa)、A (%)
    """
    import warnings
    valid = df[['应力_MPa', '应变_%']].dropna()
    stress = valid['应力_MPa'].values.astype(float)
    strain = valid['应变_%'].values.astype(float)   # 单位：%
    n = len(stress)
    if n < 10:
        return {'E_GPa': float('nan'), 'Rp02_MPa': float('nan'),
                'Rm_MPa': float('nan'), 'A_pct': float('nan')}

    # ── Rm：最大应力 ──
    rm_idx = int(np.argmax(stress))
    Rm = float(stress[rm_idx])

    # ── A：延伸率取全列最大应变 ──
    A = float(np.max(strain))

    E_GPa = float('nan')
    Rp02  = float('nan')
    try:
        # 应变转小数（% → 1）
        strain_dec = strain / 100.0

        # ── 步骤1：确定线弹性段 ──
        # 搜索范围：应力在 5%~70% Rm 之间（跳过初始噪声和塑性区）
        # 不过滤 strain==0，改为过滤应力过小的点
        lo_stress = 0.05 * Rm
        hi_stress = 0.70 * Rm
        search_mask = (stress >= lo_stress) & (stress <= hi_stress)
        idx_search = np.where(search_mask)[0]

        best_E      = float('nan')
        best_end    = 5
        best_coeffs = None

        if len(idx_search) >= 5:
            s_all  = strain_dec[idx_search]
            st_all = stress[idx_search]

            # 逐步扩展窗口，找 R²≥0.9995 的最大线性段
            for end in range(5, len(idx_search) + 1):
                s_w  = s_all[:end]
                st_w = st_all[:end]
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore')
                    coeffs = np.polyfit(s_w, st_w, 1)
                fitted = np.polyval(coeffs, s_w)
                ss_res = np.sum((st_w - fitted) ** 2)
                ss_tot = np.sum((st_w - np.mean(st_w)) ** 2)
                r2 = 1.0 - ss_res / ss_tot if ss_tot > 1e-12 else 1.0
                if r2 >= 0.9995:
                    best_E      = coeffs[0]
                    best_end    = end
                    best_coeffs = coeffs
                else:
                    break

        if math.isnan(best_E) or best_coeffs is None:
            print(f"  ⚠ 未能找到线弹性段（R²<0.9995），尝试用前20%数据强制拟合")
            hi_idx = max(10, int(len(idx_search) * 0.2)) if len(idx_search) >= 10 else len(idx_search)
            if len(idx_search) >= 5:
                s_w  = strain_dec[idx_search[:hi_idx]]
                st_w = stress[idx_search[:hi_idx]]
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore')
                    best_coeffs = np.polyfit(s_w, st_w, 1)
                best_E   = best_coeffs[0]
                best_end = hi_idx
                print(f"  → 强制拟合 E={best_E/1000:.2f} GPa（仅供参考）")

        if not math.isnan(best_E) and best_coeffs is not None:
            E_GPa = best_E / 1000.0   # GPa
            print(f"  → 弹性段拟合点数={best_end}，E={E_GPa:.2f} GPa")

            # ── 步骤2：构造 0.2% 偏置线（ISO 6892 / ASTM E8）──
            # 弹性拟合线：σ = E·ε + b
            # 偏置线：σ_offset = E·(ε - 0.002) + b  （右移 0.2%）
            E_MPa       = best_E
            b_intercept = best_coeffs[1]
            offset      = E_MPa * (strain_dec - 0.002) + b_intercept   # MPa

            # ── 步骤3：在弹性段终点之后寻找交叉点 ──
            elastic_end_strain = strain_dec[idx_search[best_end - 1]]
            diff = stress - offset

            cross_idx = None
            for i in range(1, n):
                if strain_dec[i] <= elastic_end_strain:
                    continue
                if diff[i - 1] < 0 and diff[i] >= 0:
                    cross_idx = i
                    break

            if cross_idx is not None:
                d0, d1 = diff[cross_idx - 1], diff[cross_idx]
                t = -d0 / (d1 - d0) if abs(d1 - d0) > 1e-12 else 0.0
                Rp02 = float(stress[cross_idx - 1] + t * (stress[cross_idx] - stress[cross_idx - 1]))
            else:
                # 未找到交叉：取弹性段后 diff 绝对值最小点
                post = strain_dec >= elastic_end_strain
                if post.any():
                    Rp02 = float(stress[post][np.argmin(np.abs(diff[post]))])
                else:
                    Rp02 = float(stress[np.argmin(np.abs(diff))])

    except Exception as e:
        print(f"  ⚠ 力学性能计算出错：{e}")
        import traceback; traceback.print_exc()

    return {'E_GPa': E_GPa, 'Rp02_MPa': Rp02, 'Rm_MPa': Rm, 'A_pct': A}


# ──────────────────────────────────────────────────────────
# 10. 写出结果 xlsx（每试样一个 Sheet）
# ──────────────────────────────────────────────────────────
def write_result_xlsx(results: dict, prefix: str) -> str:
    desired = f"{prefix}_result.xlsx"
    headers = ['载荷', '行程', '引伸计（修正）', '应力', '应变']
    units   = ['kN',  'mm',  'mm',            'MPa',  '%']

    print(f"\n  各试样力学性能汇总：")
    mech_results = {}
    for sp_name, df in results.items():
        props = calc_mechanical_properties(df)
        mech_results[sp_name] = props
        E_str   = f"{props['E_GPa']:.1f} GPa"   if not math.isnan(props['E_GPa'])   else "N/A"
        Rp_str  = f"{props['Rp02_MPa']:.1f} MPa" if not math.isnan(props['Rp02_MPa']) else "N/A"
        Rm_str  = f"{props['Rm_MPa']:.1f} MPa"   if not math.isnan(props['Rm_MPa'])  else "N/A"
        A_str   = f"{props['A_pct']:.2f} %"      if not math.isnan(props['A_pct'])   else "N/A"
        print(f"    {sp_name}：屈服强度={Rp_str}，抗拉强度={Rm_str}，延伸率={A_str}，杨氏模量={E_str}")

    def _do_write(out_name):
        with pd.ExcelWriter(out_name, engine='openpyxl') as writer:
            for sp_name, df in results.items():
                out_df = pd.DataFrame({
                    '载荷':        df['载荷_kN'],
                    '行程':        df['行程_mm'],
                    '引伸计（修正）': df['引伸计_修正'],
                    '应力':        df['应力_MPa'],
                    '应变':        df['应变_%'],
                })
                out_df.to_excel(writer, sheet_name=sp_name[:31],
                                index=False, header=False, startrow=2)
                ws = writer.sheets[sp_name[:31]]

                for ci, h in enumerate(headers, start=1):
                    c = ws.cell(row=1, column=ci, value=h)
                    c.font = Font(bold=True)
                    c.alignment = Alignment(horizontal='center')
                    c.fill = PatternFill(start_color='DDEEFF', end_color='DDEEFF',
                                         fill_type='solid')
                for ci, u in enumerate(units, start=1):
                    c = ws.cell(row=2, column=ci, value=u)
                    c.font = Font(italic=True, color='555555')
                    c.alignment = Alignment(horizontal='center')
                for ci in range(1, 6):
                    ws.column_dimensions[get_column_letter(ci)].width = 18

    def _do_write_with_summary(out_name):
        _do_write(out_name)
        # 追加"力学性能汇总" Sheet
        from openpyxl import load_workbook
        wb = load_workbook(out_name)
        sum_sheet_name = '力学性能汇总'
        if sum_sheet_name in wb.sheetnames:
            del wb[sum_sheet_name]
        ws_sum = wb.create_sheet(sum_sheet_name, 0)   # 插到最前面
        sum_headers = ['试样', '屈服强度 Rp0.2 (MPa)', '抗拉强度 Rm (MPa)',
                        '延伸率 A (%)', '杨氏模量 E (GPa)']
        for ci, h in enumerate(sum_headers, start=1):
            c = ws_sum.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.fill = PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid')
        for ri, (sp_name, props) in enumerate(mech_results.items(), start=2):
            ws_sum.cell(row=ri, column=1, value=sp_name)
            def _val(v):
                return round(v, 3) if not math.isnan(v) else 'N/A'
            ws_sum.cell(row=ri, column=2, value=_val(props['Rp02_MPa']))
            ws_sum.cell(row=ri, column=3, value=_val(props['Rm_MPa']))
            ws_sum.cell(row=ri, column=4, value=_val(props['A_pct']))
            ws_sum.cell(row=ri, column=5, value=_val(props['E_GPa']))
        for ci in range(1, 6):
            ws_sum.column_dimensions[get_column_letter(ci)].width = 22
        wb.save(out_name)

    out_name = _write_xlsx_safe(_do_write_with_summary, desired)
    print(f"\n  ✓ 结果已保存：{out_name}（{len(results)+1} 个工作表，含力学性能汇总）")
    return out_name


# ──────────────────────────────────────────────────────────
# 主流程
# ──────────────────────────────────────────────────────────
def main():
    print("\n========== 拉伸试验数据处理程序 v2.0 ==========")

    start_file = input("\n  请输入起始 CSV 文件路径：").strip().strip('"')
    if not os.path.exists(start_file):
        print(f"  ✗ 文件不存在：{start_file}")
        sys.exit(1)

    folder = os.path.dirname(os.path.abspath(start_file))
    os.chdir(folder)
    start_file = os.path.basename(start_file)
    name_only, _ = os.path.splitext(start_file)
    m = re.match(r'^(.+)-(\d+)$', name_only)
    prefix = m.group(1) if m else name_only

    # 步骤 1：查找文件
    print("\n─── 步骤 1：查找文件 ───")
    files = find_csv_files(start_file)

    # 步骤 2：读取 & 合并
    print("\n─── 步骤 2：读取并合并文件 ───")
    raw, specimens = load_and_merge_files(files)
    print(f"  检测到 {len(specimens)} 个试样：{[s[0] for s in specimens]}")

    # 步骤 3：保存 pre xlsx
    print("\n─── 步骤 3：保存 pre 原始文件 ───")
    save_pre_xlsx(raw, specimens, prefix)

    # 步骤 4：输入各试样参数
    print("\n─── 步骤 4：输入试样参数 ───")
    params = get_all_specimen_params(specimens)

    # 步骤 5：逐试样处理
    print("\n─── 步骤 5：引伸计检测 & 应力应变计算 ───")
    results = {}
    for sp_name, col_start in specimens:
        df = extract_specimen_df(raw, col_start)
        p = params[sp_name]
        df = process_extensometer(df, p['gauge'], sp_name)
        df = calculate_stress_strain(df, p['area'], p['gauge'])
        results[sp_name] = df

    # 步骤 6：写出结果
    print("\n─── 步骤 6：写出结果文件 ───")
    write_result_xlsx(results, prefix)

    print("\n========== 处理完成 ==========\n")


if __name__ == '__main__':
    main()
